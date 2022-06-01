
from openpyxl import load_workbook

class ExcelToJson():
    def __init__(self,file_name):
        self.wb = load_workbook(file_name)
        self.data = {}
        self.conversion()

    def getJson(self):
        return self.data

    def conversion(self):
        ret = []
        try:
            ws = self.wb["Sheet1"]
            headers = ("ProductName","Description", "Addl_desc_product","prod_req_doc","prod_deliverables","prod_eta","Category","Price","Tax","Inclusicve ","MOU","Image_url","Var1_Name","var1_Desc","var1_required_docs","var1_deliverables","var1_ETA","var1_Price","Var2_Name","var2_Desc","var2_required_docs","var2_deliverables","var2_ETA","var2_Price","var3_name","var3_Desc","var3_required_docs","var3_deliverables","var3_ETA","var3_Price")
            for i in range(1,len(headers)+1):

                if ws.cell(1,i).value != headers[i-1]:
                    raise Exception(f'Unknown Data Format, expected {headers[i-1]}, found {ws.cell(1,i).value}')

            for row in ws.iter_rows(min_row=2):
                fn_name = row[0].value
                if type(fn_name) is  str and len(fn_name) > 0:
                    for item in ret:
                        if fn_name == item['display_name']:
                            raise Exception(f"Duplicate product {fn_name} found")
                    ret.append({
                        "display_name": row[0].value,
                        "description": row[1].value,
                        "addl_description": row[2].value,
                        "req_doc": row[3].value,
                        "deliverables": row[4].value,
                        "eta": row[5].value,
                        "category": row[6].value,
                        "sale_val": row[7].value,
                        "tax_methods": row[8].value,
                        "tax_inclusicve": row[9].value,
                        "uom": row[10].value,
                        "image_url": row[11].value,
                        "variants":[{
                        "variant_display_name": row[12].value,
                        "variant_description": row[13].value,
                        "variant_required_docs": row[14].value,
                        "variant_deliverables": row[15].value,
                        "variant_eta": row[16].value,
                        "variant_sale_val": row[17].value
                        },
                        {
                        "variant_display_name": row[18].value,
                        "variant_description": row[19].value,
                        "variant_required_docs": row[20].value,
                        "variant_deliverables": row[21].value,
                        "variant_eta": row[22].value,
                        "variant_sale_val": row[23].value
                        },
                        {
                        "variant_display_name": row[24].value,
                        "variant_description": row[25].value,
                        "variant_required_docs": row[26].value,
                        "variant_deliverables": row[27].value,
                        "variant_eta": row[28].value,
                        "variant_sale_val": row[29].value
                        }]
                    })
        except KeyError:
            pass
        except Exception as e:
            print('Error is :', e)
        # self.data['products'] = ret
        self.data = ret






